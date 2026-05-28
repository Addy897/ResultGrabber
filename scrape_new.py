import cv2
import re
import os
import traceback
from pprint import pprint

import openpyxl
import openpyxl.worksheet.worksheet
import pandas as pd
from bs4 import BeautifulSoup
from fp.fp import FreeProxy
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import easyocr

reader = None
def initReader():
    global reader
    reader = easyocr.Reader(['en'])


def intc(value):
    if( value == None):
        return 0
    try:
        return int(str(value).strip())
    except TypeError:
        return 0
    except ValueError:
        return value

def parse_result_page(source: str) -> dict:
    soup = BeautifulSoup(source, "html.parser")
    std: dict = {}
    table = soup.find("table")
    if table:
        for row in table.find_all("tr"):
            cols = row.find_all("td")
            if len(cols) >= 2:
                key   = cols[0].get_text(strip=True)
                value = cols[1].get_text(strip=True).lstrip(":").strip()
                print(key,value)
                if "University Seat Number" in key:
                    std["usn"] = value
                elif "Student Name" in key:
                    std["name"] = value
    semesters: dict[int, dict] = {}
    for block in soup.find_all("div", class_="table-responsive"):
        sem_num = 0
        sem_tag = block.find(
            lambda tag: tag.name == "div"
            and re.search(r"Semester\s*:", tag.get_text(), re.IGNORECASE)
            and not tag.find("div")   
        )
        if sem_tag:
            m = re.search(r"Semester\s*:\s*(\d+)", sem_tag.get_text(), re.IGNORECASE)
            if m:
                sem_num = int(m.group(1))

        body = block.find("div", class_="divTableBody")
        if not body:
            continue
        for row in body.find_all("div", class_="divTableRow"):
            cells = row.find_all("div", class_="divTableCell")
            if len(cells) != 2:
                break
            key   = cells[0].get_text(strip=True)
            value = cells[1].get_text(strip=True).lstrip(":").strip()
            if "University Seat Number" in key:
                std["usn"] = value
            elif "Student Name" in key:
                std["name"] = value

         
        for row in body.find_all("div", class_="divTableRow")[1:]:  # skip header
            cells = row.find_all("div", class_="divTableCell")
            if len(cells) != 7:
                continue
            subject_code = cells[0].get_text(strip=True)
            semesters.setdefault(sem_num, {})[subject_code] = {
                "name":   cells[1].get_text(strip=True),
                "CIE":    intc(cells[2].get_text(strip=True)),
                "SEE":    intc(cells[3].get_text(strip=True)),
                "Total":  intc(cells[4].get_text(strip=True)),
                "Result": cells[5].get_text(strip=True),
                "date":   cells[6].get_text(strip=True),
            }

    std["semesters"] = semesters
    return std


def get_all_marks(student: dict) -> dict:
    flat: dict = {}
    for sem_marks in student["semesters"].values():
        flat.update(sem_marks)
    return flat


def get_subject_marks(student: dict, subject_code: str):
    marks = get_all_marks(student)
    info = marks.get(subject_code)
    if info:
        return info["SEE"], info["CIE"], info["Total"], info["Result"]
    return "NC", "NC", "NC", "NC"

def swqa(sheet_obj: openpyxl.worksheet.worksheet.Worksheet, col):
    row    = sheet_obj.max_row
    column = sheet_obj.max_column
    ncol   = ["Subject Code", "No. of Students", "Appeared", "Absent", "Pass", "Fail", "% Pass"]
    wrow   = row + 4
    for k, v in enumerate(col):
        sheet_obj.cell(row=wrow + 1 + k, column=2).value = v
    for k, v in enumerate(ncol):
        sheet_obj.cell(row=wrow, column=2 + k).value = v
    for i in range(2, row + 1):
        for k, j in enumerate(range(6, column + 1, 4)):
            cell_obj             = sheet_obj.cell(row=i, column=j)
            cell_appeared        = sheet_obj.cell(row=wrow + 1 + k, column=4)
            cell_total           = sheet_obj.cell(row=wrow + 1 + k, column=3)
            cellP                = sheet_obj.cell(row=wrow + 1 + k, column=6)
            cellA                = sheet_obj.cell(row=wrow + 1 + k, column=5)
            cellF                = sheet_obj.cell(row=wrow + 1 + k, column=7)
            cellPer              = sheet_obj.cell(row=wrow + 1 + k, column=8)
            cell_appeared.value  = intc(cell_appeared.value)
            cell_total.value     = intc(cell_total.value)
            cellA.value          = intc(cellA.value)
            cellP.value          = intc(cellP.value)
            cellF.value          = intc(cellF.value)
            cell_appeared.value += cell_obj.value not in ("NC", "A")
            cell_total.value    += cell_obj.value != "NC"
            cellP.value         += cell_obj.value == "P"
            cellA.value         += cell_obj.value == "A"
            cellF.value         += cell_obj.value in ("F", "X")
            per                  = (cellP.value / (cell_total.value or 1)) * 100
            cellPer.value        = float(f"{per:.2f}")


def swqal(sheet_obj: openpyxl.worksheet.worksheet.Worksheet, col, students_row: int):
    max_row    = sheet_obj.max_row
    max_column = sheet_obj.max_column
    ncol       = ["Subject Code", "FCD", "FC", "SC", "Pass", "Total Pass"]
    wrow       = max_row + 4
    for k, v in enumerate(col):
        sheet_obj.cell(row=wrow + 1 + k, column=2).value = v
    for k, v in enumerate(ncol):
        sheet_obj.cell(row=wrow, column=2 + k).value = v
    for i in range(2, students_row + 1):
        for k, j in enumerate(range(6, max_column + 1, 4)):
            cell_res   = sheet_obj.cell(row=i, column=j)
            cell_marks = sheet_obj.cell(row=i, column=j - 1)
            cell_fcd   = sheet_obj.cell(row=wrow + 1 + k, column=3)
            cell_fc    = sheet_obj.cell(row=wrow + 1 + k, column=4)
            cell_sc    = sheet_obj.cell(row=wrow + 1 + k, column=5)
            cell_p     = sheet_obj.cell(row=wrow + 1 + k, column=6)
            cell_tp    = sheet_obj.cell(row=wrow + 1 + k, column=7)
            for c in (cell_fcd, cell_fc, cell_sc, cell_p, cell_tp):
                c.value = intc(c.value)
            if cell_res.value == "P":
                cell_tp.value += 1
                m = cell_marks.value
                if   m >= 70: cell_fcd.value += 1
                elif m >= 60: cell_fc.value  += 1
                elif m >= 50: cell_sc.value  += 1
                elif m >= 40: cell_p.value   += 1


def rwqa(sheet_obj: openpyxl.worksheet.worksheet.Worksheet, students: int):
    row    = sheet_obj.max_row
    column = sheet_obj.max_column
    col    = ["Total No. of Students", "Absent", "Pass", "Fail", "% Pass"]
    wrow   = row + 4
    for k, v in enumerate(col):
        sheet_obj.cell(row=wrow, column=2 + k).value = v
    for i in range(2, students + 1):
        A, P, F = 0, 1, 0
        for j in range(6, column + 1, 4):
            v = sheet_obj.cell(row=i, column=j).value
            if   v == "A": P = 0; A = 1
            elif v == "F": F = 1; P = 0
        cell4 = sheet_obj.cell(row=wrow + 1, column=2)
        cell  = sheet_obj.cell(row=wrow + 1, column=3)
        cell1 = sheet_obj.cell(row=wrow + 1, column=4)
        cell2 = sheet_obj.cell(row=wrow + 1, column=5)
        cell3 = sheet_obj.cell(row=wrow + 1, column=6)
        for c in (cell4, cell, cell1, cell2, cell3):
            c.value = intc(c.value)
        cell4.value += 1
        cell.value  += A
        cell1.value += P
        cell2.value += F
        cell3.value  = float(f"{(cell1.value / cell4.value) * 100:.2f}")


def analysis(fname: str, students: int, col):
    wb = openpyxl.load_workbook(fname)
    ws = wb.active
    swqa(ws, col)
    rwqa(ws, students)
    swqal(ws, col, students)
    wb.save(fname)


def dump(fname: str, students: list, analyze: bool = True):
    all_codes: set = set()
    for student in students:
        all_codes.update(get_all_marks(student).keys())
    all_codes = sorted(all_codes)

    data = []
    for student in students:
        row = [student.get("name", ""), student.get("usn", "")]
        for code in all_codes:
            see, cie, total, result = get_subject_marks(student, code)
            row.extend([see, cie, total, result])
        data.append(row)

    column_names = ["Student Name", "USN"]
    for code in all_codes:
        column_names.extend([f"{code}_SEE", f"{code}_CIE", f"{code}_Total", f"{code}_Result"])

    df = pd.DataFrame(data, columns=column_names)

    os.makedirs("output", exist_ok=True)
    output_path = os.path.join("output", f"{fname}.xlsx")
    df.to_excel(output_path, index=False)
    if analyze:
        analysis(output_path, len(students) + 1, all_codes)
    return output_path


def fetch(l, url, cancel, show, use_proxy=None, sem=None):
    students = []
    opt = webdriver.ChromeOptions()
    if not show:
        opt.add_argument("--headless")
    opt.set_capability("unhandledPromptBehavior", "accept")
    if use_proxy:
        proxy = FreeProxy().get()
        opt.add_argument("--proxy-server=" + proxy)
    driver = webdriver.Chrome(options=opt)
    if(type(sem) == str):
         try:
             sem = int(sem)
         except:
             sem = None
    for n, ryb, s in l:
        for u in range(s, n + 1):
            fetched   = False
            logged_in = False

            if cancel and cancel.is_set():
                driver.quit()
                return students

            while not fetched:
                if not logged_in:
                    if cancel and cancel.is_set():
                        driver.quit()
                        return students
                    try:
                        usn = f"{ryb}{u:03d}"
                        driver.get(url)

                        body_panel = driver.find_element(By.ID, "raj")
                        img_element = body_panel.find_element(By.TAG_NAME, "img")
                        img_element.screenshot("captcha.png")

                        img = cv2.imread("captcha.png")
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                        img = cv2.medianBlur(img, 3)
                        img = cv2.threshold(img, 0, 255, cv2.THRESH_OTSU)[1]
                        img = cv2.threshold(img, 250, 255, cv2.THRESH_BINARY)[1]
                        cv2.waitKey(0)

                        try:
                            if not reader:
                                exit()
                            code = reader.readtext(img)[0][1]
                        except IndexError:
                            continue

                        code = code.replace("\n", "").replace(" ", "")
                        if not code.isalnum() or len(code) != 6:
                            continue

                        driver.find_element("name", "lns").send_keys(usn)
                        driver.find_element("name", "captchacode").send_keys(code)
                        driver.find_element("id", "submit").click()

                        try:
                            alert = WebDriverWait(driver, 0.1).until(EC.alert_is_present())
                            if alert.text == "University Seat Number is not available or Invalid..!":
                                fetched = True
                            alert.accept()
                            continue
                        except Exception:
                            try:
                                driver.find_element("id", "submit")
                            except Exception:
                                logged_in = True
                                continue

                    except Exception as e:
                        print(f"[-] Login error: {e}")

                else:
                    try:
                        source = driver.page_source
                        std    = parse_result_page(source)
                        if sem is not None:
                            std["semesters"] = {
                                k: v for k, v in std["semesters"].items() if k == sem
                            }

                        if std.get("usn") and std.get("semesters"):
                            students.append(std)
                        else:
                            pprint(std)
                    except Exception:
                        if driver.current_url == url:
                            logged_in = False
                        print(traceback.format_exc())
                        driver.refresh()
                        continue

                    fetched = True

    driver.quit()
    return students


if __name__ == "__main__":
    initReader()
    results = fetch(
        [(65, "1BI22IC", 1),(405, "1BI23IC", 400)],
        "https://results.vtu.ac.in/MJ26cbcs/index.php",
        None,
        True,
        None,
	8
    )
    pprint(results)
    if results:
        dump("8Sem", results)
