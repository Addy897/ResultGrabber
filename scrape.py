from tkinter import TRUE
import cv2
import time
from PIL import Image
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.common.exceptions import UnexpectedAlertPresentException
import pandas as pd
#from fp.fp import FreeProxy
import openpyxl,re
import openpyxl.worksheet
import openpyxl.worksheet.worksheet
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import easyocr
import os
reader =None
def initReader():
    global reader
    reader = easyocr.Reader(['en'])
def intc(value):
    try:
        return int(value)
    except TypeError:
        return 0
    except ValueError:
        return value
def swqa(sheet_obj:openpyxl.worksheet.worksheet.Worksheet,col):
    row = sheet_obj.max_row
    column = sheet_obj.max_column
    ncol=["Subject Code","No. of Students","Appeared","Absent","Pass","Fail","% Pass"]
    wrow=row+4
    for k,v in enumerate(col):
        cell=sheet_obj.cell(row=wrow+1+k,column=2)
        cell.value=v
    for k,v in enumerate(ncol):
        cell=sheet_obj.cell(row=wrow,column=2+k)
        cell.value=v
    for i in range(2,row+1):
        k=0
        for j in range(6,column+1,4):
            cell_obj = sheet_obj.cell(row=i, column=j)
            cell1 = sheet_obj.cell(row=wrow+1+k, column=4)
            cell2 = sheet_obj.cell(row=wrow+1+k, column=3)
            cellP = sheet_obj.cell(row=wrow+1+k, column=6)
            cellA = sheet_obj.cell(row=wrow+1+k, column=5)
            cellF = sheet_obj.cell(row=wrow+1+k, column=7)
            cellPer = sheet_obj.cell(row=wrow+1+k, column=8)
            cell1.value=intc(cell1.value)
            cell2.value=intc(cell2.value)
            cellA.value=intc(cellA.value)
            cellP.value=intc(cellP.value)
            cellF.value=intc(cellF.value)
            cell1.value+=1
            cell2.value+=1
            cellP.value+=cell_obj.value=="P"
            cellA.value+=cell_obj.value=="A"
            cell1.value-=cell_obj.value=="A"
            cellF.value+=cell_obj.value=="F" or cell_obj.value=="X"
            per=(cellP.value/cell2.value)*100
            cellPer.value=float(f"{per:0.2f}")
            k+=1
def swqal(sheet_obj:openpyxl.worksheet.worksheet.Worksheet,col,students_row):
    max_row = sheet_obj.max_row
    max_column = sheet_obj.max_column
    ncol=["Subject Code","FCD","FC","SC","Pass","Total Pass"]
    wrow=max_row+4
    for k,v in enumerate(col):
        cell=sheet_obj.cell(row=wrow+1+k,column=2)
        cell.value=v
    for k,v in enumerate(ncol):
        cell=sheet_obj.cell(row=wrow,column=2+k)
        cell.value=v
    for i in range(2,students_row+1):
        k=0
        for j in range(6,max_column+1,4):
            cell_res = sheet_obj.cell(row=i, column=j)
            cell_marks = sheet_obj.cell(row=i, column=j-1)
            cell_fcd = sheet_obj.cell(row=wrow+1+k, column=3)
            cell_fc = sheet_obj.cell(row=wrow+1+k, column=4)
            cell_sc = sheet_obj.cell(row=wrow+1+k, column=5)
            cell_p = sheet_obj.cell(row=wrow+1+k, column=6)
            cell_tp = sheet_obj.cell(row=wrow+1+k, column=7)
            cell_fcd.value=intc(cell_fcd.value)
            cell_fc.value=intc(cell_fc.value)
            cell_sc.value=intc(cell_sc.value)
            cell_p.value=intc(cell_p.value)
            cell_tp.value=intc(cell_tp.value)
            if(cell_res.value=="P"):
                cell_tp.value+=1
                if(cell_marks.value>=70):
                    cell_fcd.value+=1
                elif(cell_marks.value>=60):
                    cell_fc.value+=1
                elif(cell_marks.value>=50):
                    cell_sc.value+=1
                elif(cell_marks.value>=40):
                    cell_p.value+=1



            k+=1
def rwqa(sheet_obj:openpyxl.worksheet.worksheet.Worksheet,students:int):
    row = sheet_obj.max_row
    column = sheet_obj.max_column
    col=["Total No. of Students","Absent","Pass","Fail","% Pass"]
    wrow=row+4
    for k,v in enumerate(col):
        cell=sheet_obj.cell(row=wrow,column=2+k)
        cell.value=v
    for i in range(2,students+1):
        A=0
        P=1
        F=0
        for j in range(6,column+1,4):
            cell_obj = sheet_obj.cell(row=i, column=j)
            if(cell_obj.value=="A"):
                P=0
                A=1
            elif(cell_obj.value=="F"):
                F=1
                P=0

        cell=sheet_obj.cell(row=wrow+1,column=3)
        cell1=sheet_obj.cell(row=wrow+1,column=4)
        cell2=sheet_obj.cell(row=wrow+1,column=5)
        cell3=sheet_obj.cell(row=wrow+1,column=6)
        cell4=sheet_obj.cell(row=wrow+1,column=2)
        cell.value=intc(cell.value)
        cell1.value=intc(cell1.value)
        cell2.value=intc(cell2.value)
        cell3.value=intc(cell3.value)
        cell4.value=intc(cell4.value)
        cell4.value+=1
        cell.value+=A
        cell1.value+=P
        cell2.value+=F
        per=(cell1.value/cell4.value)*100
        cell3.value=float(f"{per:0.2f}")



def analysis(fname,students,col):
    wb_obj = openpyxl.load_workbook(fname)
    sheet_obj = wb_obj.active
    swqa(sheet_obj,col)
    rwqa(sheet_obj=sheet_obj,students=students)
    swqal(sheet_obj,col,students)
    wb_obj.save(fname)


def fetch(l,url,cancel,show):
    students=[]
    opt = webdriver.ChromeOptions()
    opt.add_argument("--window-size=1051,798")
    if(not show.get()):
        opt.add_argument("--headless")
    opt.set_capability('unhandledPromptBehavior', 'accept')
    #proxy = FreeProxy().get()
    #opt.add_argument('--proxy-server='+proxy)
    driver = webdriver.Chrome(options=opt)

    for n,ryb,s in l:
        for u in range(s,n+1):
            fetched = False
            logged_in=False
            if cancel and cancel.is_set():
                driver.quit()
                return students
            while not fetched:
                if(not logged_in):
                    if cancel and cancel.is_set():
                        driver.quit()
                        return students
                    try:
                        usn = f'{ryb}{u:03d}'
                        driver.get(url)
                        driver.save_screenshot('captcha.png')
                        img = cv2.imread("captcha.png")
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                        img = cv2.medianBlur(img, 3)
                        img = cv2.threshold(img, 0, 255,  cv2.THRESH_OTSU)[1]
                        img = cv2.threshold(img, 250, 255, cv2.THRESH_BINARY)[1]
                        img = img[407:547, 607:845]
                        cv2.imwrite('captcha.png',img)
                        cv2.waitKey(0)
                        try:
                            if(not reader):
                                exit()
                            code  = reader.readtext(img)[0][1]
                        except IndexError:
                            continue
                        code = code.replace("\n","").replace(" ","")
                        if(not code.isalnum() and len(code)!=6):
                            continue

                        driver.find_element("name","lns").send_keys(usn)
                        driver.find_element("name","captchacode").send_keys(code)
                        driver.find_element("id","submit").click()
                        source=driver.page_source
                        try:
                            er= WebDriverWait(driver, 0.1).until(EC.alert_is_present())
                            if(er.text=="University Seat Number is not available or Invalid..!"):
                                fetched=True

                            er.accept()
                            continue
                        except:
                            try:
                                 driver.find_element("id","submit")
                            except:
                                logged_in=True
                                continue
                    except Exception as e:
                        print(f"[-] Not logged in {e}")
                else:
                    driver.refresh()
                    source=driver.page_source
                    try:
                        soup = BeautifulSoup(source)
                        std = {}
                        table = soup.find('table')
                        rows = table.find_all('tr')
                        for row in rows:
                            cols = row.find_all('td')
                            if len(cols) > 1:
                                key = cols[0].text.strip()
                                value = cols[1].text.strip()
                                std[key] = value.replace(": ","")
                        marks = {}
                        result_table = soup.find('div', class_='divTableBody')
                        rows = result_table.find_all('div', class_='divTableRow')


                        for row in rows[1:]:
                            cells = row.find_all('div', class_='divTableCell')
                            if len(cells) == 7:
                                subject_code=cells[0].get_text(strip=True)
                                subject_info = {
                                    'CIE': intc(cells[2].get_text(strip=True)),
                                    'SEE': intc(cells[3].get_text(strip=True)),
                                    'Total': intc(cells[4].get_text(strip=True)),
                                    'Result': cells[5].get_text(strip=True),
                                }
                                marks[subject_code]=subject_info

                        std['marks']=marks
                        students.append(std)

                    except Exception as e:
                        continue
                    fetched=True

    driver.quit()
    return students

def get_subject_marks(student, subject_code):
    subject_data = student['marks'].get(subject_code)
    if subject_data:
        return subject_data['SEE'], subject_data['CIE'], subject_data['Total'], subject_data['Result']
    else:
        return None, None, None, None

def dump(fname, students,analyize=True):
    data = []
    for student in students:
        student_name = student['Student Name']
        usn = student['University Seat Number']

        subject_marks = []
        for subject_code in student['marks']:
            see, cie, total, result = get_subject_marks(student, subject_code)
            subject_marks.extend([see, cie, total, result])

        data.append([student_name, usn] + subject_marks)

    column_names = ['Student Name', 'USN']
    for subject_code in students[0]['marks']:
        column_names.extend([subject_code + '_SEE', subject_code + '_CIE', subject_code + '_Total', subject_code + '_Result'])
    df = pd.DataFrame(data, columns=column_names)
    if(not os.path.isdir('./output')):
        os.mkdir("output")
    try:
        df.to_excel(f"output/{fname}", index=False)
    except ValueError:
        fname=f"{fname}.xlsx"
        df.to_excel(f"output/{fname}",index=False)
    if(not analyize):
        return
    analysis(fname,len(students)+1,students[0]['marks'])
